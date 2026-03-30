from scrapling.fetchers import StealthyFetcher
import time
import datetime
from pathlib import Path
import re
import html
import unicodedata
import logging
from dataclasses import dataclass, field
from typing import Dict, List, Optional
from urllib.parse import urlencode

from openpyxl import load_workbook
from openpyxl.styles import Font


TEMPLATE_PATH = Path("Template") / "DanhSachGoiThau.xlsx"
RESULT_PATH = Path("Result")
BASE_URL = "https://dauthau.asia"
SEARCH_PATH = "/thongbao/moithau/"

SEARCH_PARAMS_BASE = [
    ("type_search", "1"),
    ("type_info", "1"),
    ("type_info3", "1"),
    ("ketqua_luachon_tochuc_dgts", "0"),
    ("is_advance", "0"),
    ("is_province", "0"),
    ("is_kqlcnt", "0"),
    ("type_choose_id", "0"),
    ("search_idprovincekq", "1"),
    ("search_idprovince_khtt", "1"),
    ("oda", "0"),
    ("goods_2", "0"),
    ("searchkind", "0"),
    ("type_view_open", "0"),
    ("sl_nhathau", "0"),
    ("sl_nhathau_cgtt", "0"),
    ("search_idprovince", "1"),
    ("type_org", "1"),
    ("goods", "0"),
    ("cat", "0"),
    ("search_keyword_id_province", "1"),
    ("search_devprovince", "1"),
    ("khlcnt", "0"),
    ("search_rq_province", "-1"),
    ("search_rq_province", "1"),
    ("rq_form_value", "0"),
    ("searching", "1"),
]

logger = logging.getLogger(__name__)


@dataclass
class CrawlerConfig:
    keyword: str = "số hóa"
    max_page: int = 27
    lookback_days: int = 365
    request_delay: float = 0
    max_retry: int = 2
    headless: bool = False
    end_date: datetime.date = field(default_factory=lambda: datetime.date.today())

    @property
    def start_date(self) -> datetime.date:
        return self.end_date - datetime.timedelta(days=self.lookback_days)

    @property
    def valid_years(self) -> set:
        return set(range(self.start_date.year, self.end_date.year + 1))


@dataclass
class Tender:
    ma_tbmt: str
    goi_thau: str
    ngay_dang_tai: str
    dong_thau: str
    duong_dan: str
    details: Dict[str, str] = field(default_factory=dict)

    def to_list_record(self) -> Dict[str, str]:
        return {
            "Mã TBMT": self.ma_tbmt,
            "Gói thầu": self.goi_thau,
            "Ngày đăng tải": self.ngay_dang_tai,
            "Đóng thầu": self.dong_thau,
            "Đường dẫn": self.duong_dan,
        }

    def to_export_record(self) -> Dict[str, str]:
        base = self.to_list_record()
        base.update(self.details)
        return base

# Template hiện tại dùng nhiều header viết tắt/thiếu ký tự tiếng Việt.
LIST_HEADER_ALIASES = {
    "Mã TBMT": ["M TBMT"],
    "Gói thầu": ["Gi thầu"],
    "Ngày đăng tải": ["Ngy đăng tải"],
    "Đóng thầu": ["Đng thầu", "Thời điểm đng thầu"],
}

DETAIL_HEADER_ALIASES = {
    "STT": [],
    "Mã TBMT": ["M TBMT"],
    "Ngày đăng tải": ["Ngy đăng tải"],
    "Kế hoạch": [],
    "Trạng thái gói thầu": ["Trạng thi gi thầu"],
    "Tên dự án": ["Tn dự n"],
    "Tên gói thầu": ["Tn gi thầu"],
    "Chủ đầu tư": [],
    "Mã KHLCNT": ["M KHLCNT"],
    "Tên KHLCNT": ["Tn KHLCNT"],
    "Phân loại KHLCNT": ["Phn loại KHLCNT"],
    "Trong nước/Quốc tế": [],
    "Phương thức lựa chọn nhà thầu": ["Phương thức lựa chọn nh thầu"],
    "Thời gian thực hiện hợp đồng": [],
    "Hình thức LCNT": ["Hnh thức LCNT"],
    "Thực hiện tại": [],
    "Các thông báo liên quan": ["Cc thng bo lin quan"],
    "Thời điểm đóng thầu": ["Thời điểm đng thầu"],
    "Lĩnh vực AI phân loại": ["Lĩnh vực AI phn loại"],
    "Ngành nghề AI phân loại": ["Ngnh nghề AI phn loại"],
    "Số quyết định phê duyệt": ["Số quyết định ph duyệt"],
    "Ngày phê duyệt": ["Ngy ph duyệt"],
    "Cơ quan ra quyết định phê duyệt": ["Cơ quan ra quyết định ph duyệt"],
    "Quyết định phê duyệt": ["Quyết định ph duyệt"],
    "Hình thức dự thầu": ["Hnh thức dự thầu"],
    "Nhận HSDT từ": [],
    "Chi phí nộp E-HSDT": ["Chi ph nộp E-HSDT"],
    "Địa điểm nhận E-HSDT": [],
    "Địa điểm nhận": ["Địa điểm nhận E-HSDT"],
    "E-HSDT": ["Nhận HSDT từ", "Chi phí nộp E-HSDT", "Địa điểm nhận E-HSDT"],
    "Thời điểm mở thầu": [],
    "Địa điểm mở thầu": [],
    "Giá gói thầu": ["Gi gi thầu"],
    "Bằng chữ": [],
    "Kết quả lựa chọn nhà thầu": ["Kết quả lựa chọn nh thầu", "Tên nhà thầu"],
    "Tên nhà thầu": ["Kết quả lựa chọn nhà thầu", "Kết quả lựa chọn nh thầu"],
    "Giá trúng thầu": ["Gi trng thầu", "Gi trng thầu (VND)", "Giá trúng thầu (VND)"],
    "Hình thức đảm bảo dự thầu": ["Hnh thức đảm bảo dự thầu"],
    "Thời hạn đảm bảo": ["Thời hạn đảm bảo"],
}


def _lookup_col(exact_map, norm_map, key, aliases=None):
    key_text = str(key).strip()
    col = exact_map.get(key_text)
    if col:
        return col
    col = norm_map.get(_normalize_key(key_text))
    if col:
        return col

    for alias in aliases or []:
        alias_text = str(alias).strip()
        col = exact_map.get(alias_text)
        if col:
            return col
        col = norm_map.get(_normalize_key(alias_text))
        if col:
            return col
    return None


def _normalize_key(text):
    if not text:
        return ""
    return " ".join(str(text).replace(":", " ").split()).strip()


def _extract_text(node):
    """Lấy text hiển thị của node theo cách chịu lỗi với nhiều cấu trúc HTML."""
    if not node:
        return ""

    direct_text = getattr(node, "text", "")
    if direct_text:
        # Ưu tiên text trực tiếp để tránh lặp khi node và node con chứa cùng nội dung.
        direct_value = " ".join(str(direct_text).split()).strip()
        # Một số node có text trực tiếp là "-" nhưng thực tế dữ liệu nằm ở node con.
        if direct_value and direct_value not in {"-", "--", "---"}:
            return direct_value

    parts = []
    try:
        for child in node.css("*"):
            child_text = getattr(child, "text", "")
            if child_text:
                parts.append(str(child_text))
    except Exception:
        pass

    merged = " ".join(" ".join(parts).split()).strip()
    if merged:
        return merged

    # Fallback cho trường hợp text nằm xen giữa các tag (ví dụ button có SVG + text).
    html_content = getattr(node, "html_content", "")
    if html_content:
        raw_html = str(html_content)
        raw_html = re.sub(r"<svg\b[\s\S]*?</svg>", " ", raw_html, flags=re.IGNORECASE)
        raw_html = re.sub(r"<[^>]+>", " ", raw_html)
        plain = html.unescape(raw_html)
        return " ".join(plain.split()).strip()

    return ""


def _extract_ma_tbmt(*values):
    pattern = r"IB\d+(?:\.\d+)?(?:-\d+)?"
    for val in values:
        if not val:
            continue
        match = re.search(pattern, str(val), flags=re.IGNORECASE)
        if match:
            return match.group(0).upper()
    return ""


def _is_adv_placeholder(value):
    text = (value or "").strip().lower()
    if not text:
        return False
    # Nhận diện nội dung khóa điểm/tài khoản thay vì giá trị thật.
    placeholders = [
        "click",
        "xem",
        "thong tin",
        "thông tin",
        "dang nhap",
        "đăng nhập",
        "dang ky",
        "đăng ký",
        "bi tru",
        "bị trừ",
    ]
    return ("click" in text and "xem" in text) or any(token in text for token in placeholders[4:])


def _extract_adv_field_value(page, field_id):
    nodes = page.css(f"#{field_id}")
    if not nodes:
        return ""
    value = _extract_text(nodes[0])
    if _is_adv_placeholder(value):
        return ""
    return value


def _resolve_adv_detail_field(fetcher, detail_url, field_id):
    if not fetcher or not detail_url:
        return ""

    selector = f"#{field_id} a[onclick*='click_view_detail_adv']"

    def _click_action(page):
        try:
            page.on("dialog", lambda dialog: dialog.accept())
        except Exception:
            pass
        try:
            page.wait_for_selector(selector, timeout=5000)
            page.click(selector)
            page.wait_for_timeout(0)
        except Exception:
            pass

    try:
        clicked_page = fetcher.fetch(detail_url, wait=0, page_action=_click_action)
    except Exception:
        return ""

    return _extract_adv_field_value(clicked_page, field_id)


def _extract_latest_related_notice(detail_page):
    """Lấy thông báo liên quan mới nhất theo thời điểm tìm thấy trong text/title."""
    date_patterns = [
        r"(\d{1,2}/\d{1,2}/\d{4})(?:\s+(\d{1,2}:\d{2}))?",
        r"(\d{4}-\d{1,2}-\d{1,2})(?:\s+(\d{1,2}:\d{2}))?",
    ]

    def _parse_dt(text):
        if not text:
            return None
        source = str(text)
        for pattern in date_patterns:
            match = re.search(pattern, source)
            if not match:
                continue
            date_part = match.group(1)
            time_part = match.group(2) or "00:00"
            for fmt in ("%d/%m/%Y %H:%M", "%Y-%m-%d %H:%M"):
                try:
                    return datetime.datetime.strptime(f"{date_part} {time_part}", fmt)
                except ValueError:
                    continue
        return None

    def _is_related_notice(text, href):
        haystack = f"{text or ''} {href or ''}".lower()
        keywords = [
            "thong bao", "thông báo", "tbmt", "moi thau", "mời thầu",
            "lam ro", "làm rõ", "gia han", "gia hạn", "dinh chinh", "đính chính",
        ]
        return any(k in haystack for k in keywords)

    related_candidates = []
    for node in detail_page.css("a"):
        href = node.attrib.get("href", "") if hasattr(node, "attrib") else ""
        text = _extract_text(node)
        if not href and not text:
            continue
        if not _is_related_notice(text, href):
            continue
        dt = _parse_dt(f"{text} {node.attrib.get('title', '') if hasattr(node, 'attrib') else ''}")
        display = text or href
        related_candidates.append((dt, display))

    if not related_candidates:
        return ""

    related_candidates.sort(key=lambda x: x[0] or datetime.datetime.min, reverse=True)
    return related_candidates[0][1]


def _to_absolute_url(href):
    if not href:
        return ""
    href = str(href).strip()
    if href.startswith("http://") or href.startswith("https://"):
        return href
    if href.startswith("//"):
        return f"https:{href}"
    if href.startswith("/"):
        return f"{BASE_URL}{href}"
    return f"{BASE_URL}/{href.lstrip('/')}"


def _extract_kqlcnt_url(value_node):
    if not value_node:
        return ""

    for a_node in value_node.css('a[href^="/ketqua/"], a[href*="/ketqua/"]'):
        href = a_node.attrib.get("href", "") if hasattr(a_node, "attrib") else ""
        if "/ketqua/" in str(href).lower():
            return _to_absolute_url(href)

    node_text = _extract_text(value_node).lower()
    if "xem chi tiết" in node_text or "xem chi tiet" in node_text:
        for a_node in value_node.css("a[href]"):
            href = a_node.attrib.get("href", "") if hasattr(a_node, "attrib") else ""
            href_lower = str(href).lower()
            if not href_lower:
                continue
            if "vip" in href_lower or "bang-gia-goi-vip" in href_lower:
                continue
            return _to_absolute_url(href)

    return ""


def _parse_kqlcnt_award_page(result_page):
    def _pick_award_rows(page):
        for selector in (
            "tbody#list_business_online tr",
            "div.table-responsive.bidding_table tbody tr",
            "table tbody tr",
        ):
            rows = page.css(selector)
            if rows:
                return rows
        return []

    def _slug(text):
        base = unicodedata.normalize("NFD", str(text or ""))
        base = "".join(ch for ch in base if unicodedata.category(ch) != "Mn")
        return _normalize_key(base).lower()

    def _map_award_columns(page):
        bidder_idx = 3
        price_idx = 6
        headers = page.css("table thead tr th")
        if not headers:
            return bidder_idx, price_idx

        for idx, th in enumerate(headers):
            label = _slug(_extract_text(th))
            if "nha thau" in label and "ten" in label:
                bidder_idx = idx
            if "gia" in label and "trung" in label and "thau" in label:
                price_idx = idx
        return bidder_idx, price_idx

    def _extract_price_from_row(tds, price_idx):
        if 0 <= price_idx < len(tds):
            value = _extract_text(tds[price_idx])
            if value and value not in {"-", "--", "---"}:
                return value

        for td in tds:
            text = _extract_text(td)
            if not text:
                continue
            if re.search(r"\b(vnd|vnđ|đồng|usd|eur)\b", text.lower()) and re.search(r"\d", text):
                return text
        return ""

    def _extract_bidder_from_row(row, tds, bidder_idx):
        if 0 <= bidder_idx < len(tds):
            bidder_link_nodes = tds[bidder_idx].css("a")
            bidder_name = _extract_text(bidder_link_nodes[0]) if bidder_link_nodes else _extract_text(tds[bidder_idx])
            if bidder_name:
                return bidder_name

        # Fallback: nhiều trang đặt link nhà thầu theo class/href đặc thù.
        for a_node in row.css('a[href*="/businesslistings/detail/"]'):
            bidder_name = _extract_text(a_node)
            if bidder_name:
                return bidder_name
        return ""

    bidder_names = []
    award_prices = []
    last_known_price = ""
    bidder_idx, price_idx = _map_award_columns(result_page)

    for row in _pick_award_rows(result_page):
        tds = row.css("td")
        if not tds:
            continue

        bidder_name = _extract_bidder_from_row(row, tds, bidder_idx)

        award_price = _extract_price_from_row(tds, price_idx)
        if award_price:
            last_known_price = award_price
        elif bidder_name and last_known_price:
            # Dòng liên danh thường không có ô giá do rowspan, dùng lại giá gần nhất.
            award_price = last_known_price

        if bidder_name:
            bidder_names.append(bidder_name)
            if award_price:
                award_prices.append(award_price)

    unique_prices = []
    seen = set()
    for price in award_prices:
        norm_price = _normalize_key(price)
        if not norm_price or norm_price in seen:
            continue
        seen.add(norm_price)
        unique_prices.append(price)

    return {
        "Tên nhà thầu": "; ".join(bidder_names),
        "Giá trúng thầu": "; ".join(unique_prices),
    }


def _parse_detail_fields(detail_page, fetcher=None, detail_url=""):
    details = {}
    bidding_items = detail_page.css('.bidding-detail-item')
    for b_item in bidding_items:
        tit_nodes = b_item.css('.c-tit')
        val_nodes = b_item.css('.c-val')
        if not tit_nodes or not val_nodes:
            continue

        for idx, tit_node in enumerate(tit_nodes):
            # Key nên ưu tiên text trực tiếp (giống logic cũ) để khớp header trong template STT.
            raw_key = getattr(tit_node, "text", "")
            raw_key = str(raw_key).strip() if raw_key else ""
            if not raw_key:
                raw_key = _extract_text(tit_node).strip()
            if not raw_key:
                continue

            value_node = val_nodes[idx] if idx < len(val_nodes) else val_nodes[0]
            value = _extract_text(value_node)

            # Trích link chi tiết KQLCNT ngay tại node value để không mất href.
            if _normalize_key(raw_key) == _normalize_key("Kết quả lựa chọn nhà thầu"):
                kqlcnt_url = _extract_kqlcnt_url(value_node)
                if kqlcnt_url:
                    details["KQLCNT_URL"] = kqlcnt_url
                elif value and "chưa có kết quả" in value.lower():
                    value = "Chưa có kết quả"

            if value:
                # Giữ key gốc để khớp template chính xác, đồng thời lưu key chuẩn hóa để fallback.
                details[raw_key] = value
                norm_key = _normalize_key(raw_key)
                if norm_key and norm_key not in details:
                    details[norm_key] = value

    # Bổ sung riêng Mã TBMT vì nhiều trang dùng span.bd-code thay vì text phẳng.
    ma_nodes = detail_page.css('.bd-code, .bidding-code')
    if ma_nodes:
        details["Mã TBMT"] = _extract_ma_tbmt(_extract_text(ma_nodes[0]), details.get("Mã TBMT", ""))

    # Trường "Kế hoạch" nằm ngoài khối .bidding-detail-item ở dạng button.
    ke_hoach_nodes = detail_page.css('button.btn-dayleft, button.btn.btn-dayleft')
    if ke_hoach_nodes:
        ke_hoach_text = _extract_text(ke_hoach_nodes[0])
        if ke_hoach_text:
            details["Kế hoạch"] = ke_hoach_text

    related_notice = _extract_latest_related_notice(detail_page)
    if related_notice:
        details["Các thông báo liên quan"] = related_notice

    # Đồng bộ key để tăng khả năng map sang template mới.
    if details.get("Địa điểm nhận E-HSDT") and not details.get("Địa điểm nhận"):
        details["Địa điểm nhận"] = details["Địa điểm nhận E-HSDT"]
    if details.get("Nhận HSDT từ") and not details.get("E-HSDT"):
        details["E-HSDT"] = details["Nhận HSDT từ"]
    if details.get("Trong nước - Quốc tế") and not details.get("Trong nước/Quốc tế"):
        details["Trong nước/Quốc tế"] = details["Trong nước - Quốc tế"]

    # Nâng cấp: với trường bị khóa điểm, tự click để lấy giá trị thật khi có thể.
    # [COMMENT] Tạm thời tắt tính năng click để tối ưu tốc độ - để trường trống
    # contract_field = "Thời gian thực hiện hợp đồng"
    # contract_value = details.get(contract_field, "")
    # if not contract_value:
    #     contract_value = _extract_adv_field_value(detail_page, "thoi_gian_thuc_hien")
    #     if contract_value:
    #         details[contract_field] = contract_value
    # if _is_adv_placeholder(contract_value):
    #     resolved = _resolve_adv_detail_field(fetcher, detail_url, "thoi_gian_thuc_hien")
    #     if resolved:
    #         details[contract_field] = resolved
    #         norm_key = _normalize_key(contract_field)
    #         if norm_key and norm_key not in details:
    #             details[norm_key] = resolved
    #     else:
    #         details[contract_field] = ""
    #         norm_key = _normalize_key(contract_field)
    #         if norm_key and norm_key in details:
    #             details[norm_key] = ""

    return details

def build_search_url(page_number: int, config: CrawlerConfig) -> str:
    dynamic_params = [
        ("q", config.keyword),
        ("sfrom", config.start_date.strftime("%d/%m/%Y")),
        ("sto", config.end_date.strftime("%d/%m/%Y")),
        ("page", str(page_number)),
    ]
    query = urlencode(dynamic_params + SEARCH_PARAMS_BASE, doseq=True)
    return f"{BASE_URL}{SEARCH_PATH}?{query}"


def fetch_page(fetcher, url: str, config: CrawlerConfig, wait: int = 0):
    for attempt in range(1, config.max_retry + 2):
        try:
            return fetcher.fetch(url, wait=wait)
        except Exception as ex:
            logger.warning("Fetch lỗi (lần %s): %s | %s", attempt, url, ex)
            if attempt > config.max_retry:
                return None
            if config.request_delay:
                time.sleep(config.request_delay)
    return None


def fetch_list_page(fetcher, page_number: int, config: CrawlerConfig):
    url = build_search_url(page_number, config)
    logger.info("[TRANG %s] Truy cập: %s", page_number, url)
    return fetch_page(fetcher, url, config)


def fetch_detail_page(fetcher, detail_url: str, config: CrawlerConfig):
    return fetch_page(fetcher, detail_url, config)


def is_valid_item(goi_thau: str, ngay_dang: str, config: CrawlerConfig) -> bool:
    if not goi_thau or config.keyword.lower() not in goi_thau.lower():
        return False

    try:
        year = int(str(ngay_dang)[-4:])
    except Exception:
        return False
    return year in config.valid_years


def parse_list_page(page, config: CrawlerConfig) -> List[Tender]:
    results: List[Tender] = []
    cac_dong = page.css("tr")
    logger.debug("Số lượng hàng (tr) tìm thấy: %s", len(cac_dong))

    for dong in cac_dong:
        a_goi_thau = dong.css('td[data-column="Gói thầu"] a')
        if not a_goi_thau:
            continue

        goi_thau = a_goi_thau[0].attrib.get("title", "").strip()
        if not goi_thau:
            goi_thau = (a_goi_thau[0].text or "").strip()

        ma_tbmt_node = a_goi_thau[0].css("span.bidding-code")
        ma_tbmt = _extract_ma_tbmt(
            _extract_text(ma_tbmt_node[0]) if ma_tbmt_node else "",
            a_goi_thau[0].attrib.get("title", ""),
            a_goi_thau[0].text,
        )

        full_link = _to_absolute_url(a_goi_thau[0].attrib.get("href", ""))
        if not goi_thau:
            goi_thau = full_link.split("/")[-1]

        div_ngay_dang = dong.css('td[data-column="Ngày đăng tải"] div')
        ngay_dang = (div_ngay_dang[0].text or "").strip() if div_ngay_dang else ""

        div_dong_thau = dong.css('td[data-column="Đóng thầu"] div')
        dong_thau = _extract_text(div_dong_thau[0]) if div_dong_thau else ""

        if not is_valid_item(goi_thau, ngay_dang, config):
            continue

        tender = Tender(
            ma_tbmt=ma_tbmt,
            goi_thau=goi_thau,
            ngay_dang_tai=ngay_dang,
            dong_thau=dong_thau,
            duong_dan=full_link,
        )
        results.append(tender)
        logger.debug("[CHECK] %s | Ngày đăng: %s", goi_thau, ngay_dang)

    return results


def crawl_list(fetcher, config: CrawlerConfig) -> List[Tender]:
    filtered_links: List[Tender] = []
    for so_trang in range(1):
    # for so_trang in range(1, config.max_page + 1):

        page = fetch_list_page(fetcher, so_trang, config)
        if page is None:
            logger.error("Không thể tải trang %s sau khi retry.", so_trang)
            break

        if so_trang == 1:
            logger.info("Nếu có captcha, xử lý trên trình duyệt rồi tiếp tục.")

        if page.status != 200:
            logger.error("Lỗi trang %s (status=%s). Dừng crawler list.", so_trang, page.status)
            break

        parsed = parse_list_page(page, config)
        filtered_links.extend(parsed)
        logger.info("Hoàn thành trang %s: %s gói thầu phù hợp.", so_trang, len(parsed))

        if config.request_delay:
            time.sleep(config.request_delay)

    return filtered_links


def crawl_detail(fetcher, tender: Tender, config: CrawlerConfig) -> Optional[Tender]:
    logger.info("Lấy chi tiết: %s", tender.duong_dan)
    detail_page = fetch_detail_page(fetcher, tender.duong_dan, config)
    if detail_page is None:
        logger.warning("Bỏ qua link do lỗi fetch: %s", tender.duong_dan)
        return None
    if detail_page.status != 200:
        logger.warning("Lỗi lấy chi tiết: %s | Status: %s", tender.duong_dan, detail_page.status)
        return None

    details = _parse_detail_fields(detail_page, fetcher=fetcher, detail_url=tender.duong_dan)

    kqlcnt_url = details.get("KQLCNT_URL", "")
    if kqlcnt_url:
        kqlcnt_page = fetch_page(fetcher, kqlcnt_url, config, wait=0)
        if kqlcnt_page and kqlcnt_page.status == 200:
            kqlcnt_data = _parse_kqlcnt_award_page(kqlcnt_page)
            bidder_names = kqlcnt_data.get("Tên nhà thầu", "")
            award_prices = kqlcnt_data.get("Giá trúng thầu", "")
            if bidder_names:
                details["Kết quả lựa chọn nhà thầu"] = bidder_names
                details["Tên nhà thầu"] = bidder_names
            if award_prices:
                details["Giá trúng thầu"] = award_prices
        else:
            logger.warning("Không lấy được trang KQLCNT: %s", kqlcnt_url)

    if not details.get("Kết quả lựa chọn nhà thầu"):
        details["Kết quả lựa chọn nhà thầu"] = "Chưa có kết quả"
    if not details.get("Tên nhà thầu"):
        details["Tên nhà thầu"] = details.get("Kết quả lựa chọn nhà thầu", "")
    if not details.get("Giá trúng thầu"):
        details["Giá trúng thầu"] = "" if kqlcnt_url else "Chưa có kết quả"

    tender.ma_tbmt = _extract_ma_tbmt(details.get("Mã TBMT", ""), tender.ma_tbmt) or tender.ma_tbmt
    if not tender.dong_thau:
        tender.dong_thau = details.get("Thời điểm đóng thầu", "") or details.get("Thời điểm đng thầu", "")
    tender.details = details
    return tender


def crawler_dauthau_chuyen_nghiep(config: Optional[CrawlerConfig] = None):
    config = config or CrawlerConfig()
    logger.info("=== KHỞI ĐỘNG CRAWLER CHUYÊN NGHIỆP ===")

    fetcher = StealthyFetcher(headless=config.headless)
    list_items = crawl_list(fetcher, config)

    all_data: List[Dict[str, str]] = []
    for tender in list_items:
        tender_with_detail = crawl_detail(fetcher, tender, config)
        if tender_with_detail:
            all_data.append(tender_with_detail.to_export_record())
        if config.request_delay:
            time.sleep(config.request_delay)

    if all_data:
        logger.info("Tổng số bản ghi: %s", len(all_data))
        file_name = export_to_template_workbook(all_data)
        logger.info("Đã thu thập %s gói thầu số hóa. File: %s", len(all_data), file_name)
    else:
        logger.warning("Không lấy được dữ liệu phù hợp.")

    return [item.to_list_record() for item in list_items]

def crawl_detail_dauthau(url, config: Optional[CrawlerConfig] = None):
    config = config or CrawlerConfig(headless=True)
    fetcher = StealthyFetcher(headless=config.headless)
    page = fetch_detail_page(fetcher, url, config)
    if page and page.status == 200:
        return _parse_detail_fields(page, fetcher=fetcher, detail_url=url)
    return None


def export_to_template_workbook(all_data):
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Không tìm thấy template: {TEMPLATE_PATH}")

    wb = load_workbook(TEMPLATE_PATH)
    if "DanhSachGoiThau" not in wb.sheetnames:
        raise ValueError("Template phải có sheet 'DanhSachGoiThau'.")

    detail_sheet_name = "Detail" if "Detail" in wb.sheetnames else "STT" if "STT" in wb.sheetnames else None
    if not detail_sheet_name:
        raise ValueError("Template phải có sheet detail: 'Detail' (mới) hoặc 'STT' (cũ).")

    ws_list = wb["DanhSachGoiThau"]
    ws_detail_template = wb[detail_sheet_name]

    list_headers = {}
    list_headers_norm = {}
    for col in range(1, ws_list.max_column + 1):
        header = ws_list.cell(1, col).value
        if header:
            header_text = str(header).strip()
            list_headers[header_text] = col
            list_headers_norm[_normalize_key(header_text)] = col

    def _get_list_col(*candidates):
        for c in candidates:
            col = _lookup_col(list_headers, list_headers_norm, c, LIST_HEADER_ALIASES.get(c, []))
            if col:
                return col
        return None

    stt_col = _get_list_col("STT")
    ma_tbmt_col = _get_list_col("Mã TBMT")
    goi_thau_col = _get_list_col("Gói thầu")
    ngay_dang_col = _get_list_col("Ngày đăng tải")
    duong_dan_col = _get_list_col("Đường dẫn")
    dong_thau_col = _get_list_col("Đóng thầu", "Thời điểm đóng thầu")

    # Lập map cột của sheet detail dựa trên hàng tiêu đề (row 1).
    detail_header_col_map = {}
    detail_header_col_map_norm = {}
    for col in range(1, ws_detail_template.max_column + 1):
        header = ws_detail_template.cell(1, col).value
        if header:
            header_text = str(header).strip()
            detail_header_col_map[header_text] = col
            detail_header_col_map_norm[_normalize_key(header_text)] = col

    # Xoa du lieu cu tren sheet detail (tu dong 2 tro di) de tranh ton du lieu lan chay truoc.
    for detail_row in range(2, ws_detail_template.max_row + 1):
        for col in detail_header_col_map.values():
            ws_detail_template.cell(detail_row, col).value = ""

    for idx, record in enumerate(all_data, start=1):
        row = idx + 1
        detail_row = idx + 1
        url = record.get("Đường dẫn", "")

        # Ghi detail cua tung ban ghi vao cung 1 sheet detail, moi ban ghi 1 dong.
        stt_detail_col = _lookup_col(detail_header_col_map, detail_header_col_map_norm, "STT", DETAIL_HEADER_ALIASES.get("STT", []))
        ngay_dang_detail_col = _lookup_col(detail_header_col_map, detail_header_col_map_norm, "Ngày đăng tải", DETAIL_HEADER_ALIASES.get("Ngày đăng tải", []))
        if stt_detail_col:
            ws_detail_template.cell(detail_row, stt_detail_col).value = idx
        if ngay_dang_detail_col:
            ws_detail_template.cell(detail_row, ngay_dang_detail_col).value = record.get("Ngày đăng tải", "")

        for key, value in record.items():
            key_text = str(key).strip()
            col = _lookup_col(
                detail_header_col_map,
                detail_header_col_map_norm,
                key_text,
                DETAIL_HEADER_ALIASES.get(key_text, []),
            )
            if col:
                ws_detail_template.cell(detail_row, col).value = value

        if stt_col:
            cell = ws_list.cell(row, stt_col)
            cell.value = idx
            cell.hyperlink = f"#'{detail_sheet_name}'!A{detail_row}"
            cell.font = Font(color="0563C1", underline="single")

        if ma_tbmt_col:
            ws_list.cell(row, ma_tbmt_col).value = record.get("Mã TBMT", "")
        if goi_thau_col:
            ws_list.cell(row, goi_thau_col).value = record.get("Gói thầu", "")
        if ngay_dang_col:
            ws_list.cell(row, ngay_dang_col).value = record.get("Ngày đăng tải", "")
        if dong_thau_col:
            ws_list.cell(row, dong_thau_col).value = record.get("Đóng thầu", "")
        if duong_dan_col:
            ws_list.cell(row, duong_dan_col).value = url


    thoi_gian_hien_tai = time.strftime("%d%m%Y_%H%M%S")
    RESULT_PATH.mkdir(parents=True, exist_ok=True)
    output_path = RESULT_PATH / f"DanhSachGoiThau_SoHoa_{thoi_gian_hien_tai}.xlsx"
    wb.save(output_path)
    return str(output_path)

if __name__ == "__main__":
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(message)s",
    )
    crawler_dauthau_chuyen_nghiep()

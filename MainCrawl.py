from scrapling.fetchers import StealthyFetcher
import time
import datetime
from pathlib import Path
import re
import html

from openpyxl import load_workbook
from openpyxl.styles import Font


TEMPLATE_PATH = Path("Template") / "DanhSachGoiThau.xlsx"
RESULT_PATH = Path("Result")

# Template hiện tại dùng nhiều header viết tắt/thiếu ký tự tiếng Việt.
LIST_HEADER_ALIASES = {
    "Mã TBMT": ["M TBMT"],
    "Gói thầu": ["Gi thầu"],
    "Ngày đăng tải": ["Ngy đăng tải"],
    "Đóng thầu": ["Đng thầu", "Thời điểm đng thầu"],
}

DETAIL_HEADER_ALIASES = {
    "STT": [],
    "Mã TBMT": ["M TBMT"],
    "Ngày đăng tải": ["Ngy đăng tải"],
    "Kế hoạch": [],
    "Trạng thái gói thầu": ["Trạng thi gi thầu"],
    "Tên dự án": ["Tn dự n"],
    "Tên gói thầu": ["Tn gi thầu"],
    "Chủ đầu tư": [],
    "Mã KHLCNT": ["M KHLCNT"],
    "Tên KHLCNT": ["Tn KHLCNT"],
    "Phân loại KHLCNT": ["Phn loại KHLCNT"],
    "Trong nước/Quốc tế": [],
    "Phương thức lựa chọn nhà thầu": ["Phương thức lựa chọn nh thầu"],
    "Thời gian thực hiện hợp đồng": [],
    "Hình thức LCNT": ["Hnh thức LCNT"],
    "Thực hiện tại": [],
    "Các thông báo liên quan": ["Cc thng bo lin quan"],
    "Thời điểm đóng thầu": ["Thời điểm đng thầu"],
    "Lĩnh vực AI phân loại": ["Lĩnh vực AI phn loại"],
    "Ngành nghề AI phân loại": ["Ngnh nghề AI phn loại"],
    "Số quyết định phê duyệt": ["Số quyết định ph duyệt"],
    "Ngày phê duyệt": ["Ngy ph duyệt"],
    "Cơ quan ra quyết định phê duyệt": ["Cơ quan ra quyết định ph duyệt"],
    "Quyết định phê duyệt": ["Quyết định ph duyệt"],
    "Hình thức dự thầu": ["Hnh thức dự thầu"],
    "Nhận HSDT từ": [],
    "Chi phí nộp E-HSDT": ["Chi ph nộp E-HSDT"],
    "Địa điểm nhận E-HSDT": [],
    "Địa điểm nhận": ["Địa điểm nhận E-HSDT"],
    "E-HSDT": ["Nhận HSDT từ", "Chi phí nộp E-HSDT", "Địa điểm nhận E-HSDT"],
    "Thời điểm mở thầu": [],
    "Địa điểm mở thầu": [],
    "Giá gói thầu": ["Gi gi thầu"],
    "Bằng chữ": [],
    "Kết quả lựa chọn nhà thầu": ["Kết quả lựa chọn nh thầu"],
    "Hình thức đảm bảo dự thầu": ["Hnh thức đảm bảo dự thầu"],
    "Thời hạn đảm bảo": ["Thời hạn đảm bảo"],
}


def _lookup_col(exact_map, norm_map, key, aliases=None):
    key_text = str(key).strip()
    col = exact_map.get(key_text)
    if col:
        return col
    col = norm_map.get(_normalize_key(key_text))
    if col:
        return col

    for alias in aliases or []:
        alias_text = str(alias).strip()
        col = exact_map.get(alias_text)
        if col:
            return col
        col = norm_map.get(_normalize_key(alias_text))
        if col:
            return col
    return None


def _normalize_key(text):
    if not text:
        return ""
    return " ".join(str(text).replace(":", " ").split()).strip()


def _extract_text(node):
    """Lấy text hiển thị của node theo cách chịu lỗi với nhiều cấu trúc HTML."""
    if not node:
        return ""

    direct_text = getattr(node, "text", "")
    if direct_text:
        # Ưu tiên text trực tiếp để tránh lặp khi node và node con chứa cùng nội dung.
        direct_value = " ".join(str(direct_text).split()).strip()
        # Một số node có text trực tiếp là "-" nhưng thực tế dữ liệu nằm ở node con.
        if direct_value and direct_value not in {"-", "--", "---"}:
            return direct_value

    parts = []
    try:
        for child in node.css("*"):
            child_text = getattr(child, "text", "")
            if child_text:
                parts.append(str(child_text))
    except Exception:
        pass

    merged = " ".join(" ".join(parts).split()).strip()
    if merged:
        return merged

    # Fallback cho trường hợp text nằm xen giữa các tag (ví dụ button có SVG + text).
    html_content = getattr(node, "html_content", "")
    if html_content:
        raw_html = str(html_content)
        raw_html = re.sub(r"<svg\b[\s\S]*?</svg>", " ", raw_html, flags=re.IGNORECASE)
        raw_html = re.sub(r"<[^>]+>", " ", raw_html)
        plain = html.unescape(raw_html)
        return " ".join(plain.split()).strip()

    return ""


def _extract_ma_tbmt(*values):
    pattern = r"IB\d+(?:\.\d+)?(?:-\d+)?"
    for val in values:
        if not val:
            continue
        match = re.search(pattern, str(val), flags=re.IGNORECASE)
        if match:
            return match.group(0).upper()
    return ""


def _is_adv_placeholder(value):
    text = (value or "").strip().lower()
    if not text:
        return False
    # Nhận diện nội dung khóa điểm/tài khoản thay vì giá trị thật.
    placeholders = [
        "click",
        "xem",
        "thong tin",
        "thông tin",
        "dang nhap",
        "đăng nhập",
        "dang ky",
        "đăng ký",
        "bi tru",
        "bị trừ",
    ]
    return ("click" in text and "xem" in text) or any(token in text for token in placeholders[4:])


def _extract_adv_field_value(page, field_id):
    nodes = page.css(f"#{field_id}")
    if not nodes:
        return ""
    value = _extract_text(nodes[0])
    if _is_adv_placeholder(value):
        return ""
    return value


def _resolve_adv_detail_field(fetcher, detail_url, field_id):
    if not fetcher or not detail_url:
        return ""

    selector = f"#{field_id} a[onclick*='click_view_detail_adv']"

    def _click_action(page):
        try:
            page.on("dialog", lambda dialog: dialog.accept())
        except Exception:
            pass
        try:
            page.wait_for_selector(selector, timeout=5000)
            page.click(selector)
            page.wait_for_timeout(0)
        except Exception:
            pass

    try:
        clicked_page = fetcher.fetch(detail_url, wait=0, page_action=_click_action)
    except Exception:
        return ""

    return _extract_adv_field_value(clicked_page, field_id)


def _extract_latest_related_notice(detail_page):
    """Lấy thông báo liên quan mới nhất theo thời điểm tìm thấy trong text/title."""
    date_patterns = [
        r"(\d{1,2}/\d{1,2}/\d{4})(?:\s+(\d{1,2}:\d{2}))?",
        r"(\d{4}-\d{1,2}-\d{1,2})(?:\s+(\d{1,2}:\d{2}))?",
    ]

    def _parse_dt(text):
        if not text:
            return None
        source = str(text)
        for pattern in date_patterns:
            match = re.search(pattern, source)
            if not match:
                continue
            date_part = match.group(1)
            time_part = match.group(2) or "00:00"
            for fmt in ("%d/%m/%Y %H:%M", "%Y-%m-%d %H:%M"):
                try:
                    return datetime.datetime.strptime(f"{date_part} {time_part}", fmt)
                except ValueError:
                    continue
        return None

    def _is_related_notice(text, href):
        haystack = f"{text or ''} {href or ''}".lower()
        keywords = [
            "thong bao", "thông báo", "tbmt", "moi thau", "mời thầu",
            "lam ro", "làm rõ", "gia han", "gia hạn", "dinh chinh", "đính chính",
        ]
        return any(k in haystack for k in keywords)

    related_candidates = []
    for node in detail_page.css("a"):
        href = node.attrib.get("href", "") if hasattr(node, "attrib") else ""
        text = _extract_text(node)
        if not href and not text:
            continue
        if not _is_related_notice(text, href):
            continue
        dt = _parse_dt(f"{text} {node.attrib.get('title', '') if hasattr(node, 'attrib') else ''}")
        display = text or href
        related_candidates.append((dt, display))

    if not related_candidates:
        return ""

    related_candidates.sort(key=lambda x: x[0] or datetime.datetime.min, reverse=True)
    return related_candidates[0][1]


def _parse_detail_fields(detail_page, fetcher=None, detail_url=""):
    details = {}
    bidding_items = detail_page.css('.bidding-detail-item')
    for b_item in bidding_items:
        tit_nodes = b_item.css('.c-tit')
        val_nodes = b_item.css('.c-val')
        if not tit_nodes or not val_nodes:
            continue

        for idx, tit_node in enumerate(tit_nodes):
            # Key nên ưu tiên text trực tiếp (giống logic cũ) để khớp header trong template STT.
            raw_key = getattr(tit_node, "text", "")
            raw_key = str(raw_key).strip() if raw_key else ""
            if not raw_key:
                raw_key = _extract_text(tit_node).strip()
            if not raw_key:
                continue

            value_node = val_nodes[idx] if idx < len(val_nodes) else val_nodes[0]
            value = _extract_text(value_node)
            if value:
                # Giữ key gốc để khớp template chính xác, đồng thời lưu key chuẩn hóa để fallback.
                details[raw_key] = value
                norm_key = _normalize_key(raw_key)
                if norm_key and norm_key not in details:
                    details[norm_key] = value

    # Bổ sung riêng Mã TBMT vì nhiều trang dùng span.bd-code thay vì text phẳng.
    ma_nodes = detail_page.css('.bd-code, .bidding-code')
    if ma_nodes:
        details["Mã TBMT"] = _extract_ma_tbmt(_extract_text(ma_nodes[0]), details.get("Mã TBMT", ""))

    # Trường "Kế hoạch" nằm ngoài khối .bidding-detail-item ở dạng button.
    ke_hoach_nodes = detail_page.css('button.btn-dayleft, button.btn.btn-dayleft')
    if ke_hoach_nodes:
        ke_hoach_text = _extract_text(ke_hoach_nodes[0])
        if ke_hoach_text:
            details["Kế hoạch"] = ke_hoach_text

    related_notice = _extract_latest_related_notice(detail_page)
    if related_notice:
        details["Các thông báo liên quan"] = related_notice

    # Đồng bộ key để tăng khả năng map sang template mới.
    if details.get("Địa điểm nhận E-HSDT") and not details.get("Địa điểm nhận"):
        details["Địa điểm nhận"] = details["Địa điểm nhận E-HSDT"]
    if details.get("Nhận HSDT từ") and not details.get("E-HSDT"):
        details["E-HSDT"] = details["Nhận HSDT từ"]
    if details.get("Trong nước - Quốc tế") and not details.get("Trong nước/Quốc tế"):
        details["Trong nước/Quốc tế"] = details["Trong nước - Quốc tế"]

    # Nâng cấp: với trường bị khóa điểm, tự click để lấy giá trị thật khi có thể.
    # [COMMENT] Tạm thời tắt tính năng click để tối ưu tốc độ - để trường trống
    # contract_field = "Thời gian thực hiện hợp đồng"
    # contract_value = details.get(contract_field, "")
    # if not contract_value:
    #     contract_value = _extract_adv_field_value(detail_page, "thoi_gian_thuc_hien")
    #     if contract_value:
    #         details[contract_field] = contract_value
    # if _is_adv_placeholder(contract_value):
    #     resolved = _resolve_adv_detail_field(fetcher, detail_url, "thoi_gian_thuc_hien")
    #     if resolved:
    #         details[contract_field] = resolved
    #         norm_key = _normalize_key(contract_field)
    #         if norm_key and norm_key not in details:
    #             details[norm_key] = resolved
    #     else:
    #         details[contract_field] = ""
    #         norm_key = _normalize_key(contract_field)
    #         if norm_key and norm_key in details:
    #             details[norm_key] = ""

    return details

def crawler_dauthau_chuyen_nghiep():
    print("=== KHỞI ĐỘNG CRAWLER CHUYÊN NGHIỆP ===")

    # Khởi tạo trình duyệt, hiện giao diện để bạn dễ xử lý Captcha
    fetcher = StealthyFetcher(headless=False)
    filtered_links = []  # Danh sách link đã lọc

    # Ngày hiện tại và ngày một năm trước
    current_date = datetime.datetime(2026, 3, 26)
    one_year_ago = current_date - datetime.timedelta(days=365)

    # Cài đặt số trang muốn cào (Ví dụ: Từ trang 1 đến trang 3)
    for so_trang in range(1,28):
    # for so_trang in range(1):
        url = f"https://dauthau.asia/thongbao/moithau/?q=s%E1%BB%91+h%C3%B3a&type_search=1&type_info=1&type_info3=1&ketqua_luachon_tochuc_dgts=0&sfrom=26%2F03%2F2025&sto=26%2F03%2F2026&is_advance=0&is_province=0&is_kqlcnt=0&type_choose_id=0&search_idprovincekq=1&search_idprovince_khtt=1&oda=0&goods_2=0&searchkind=0&type_view_open=0&sl_nhathau=0&sl_nhathau_cgtt=0&search_idprovince=1&type_org=1&goods=0&cat=0&search_keyword_id_province=1&search_devprovince=1&oda=0&khlcnt=0&search_rq_province=-1&search_rq_province=1&rq_form_value=0&searching=1&page={so_trang}"
        print(f"\n---> [TRANG {so_trang}] Đang truy cập: {url}")

        # Truy cập trang và đợi load xong JavaScript
        page = fetcher.fetch(url)

        # Xử lý thời gian chờ để vượt mặt hệ thống chống Bot
        if so_trang == 1:
            print("ĐANG ĐỢI 0 GIÂY: Hãy nhấp giải Captcha trên trình duyệt (nếu có)!")
            time.sleep(0)
        else:
            print("Nghỉ ngơi 0 giây để tránh bị khóa IP...")
            time.sleep(0)

        # Nếu vào trang thành công
        if page.status == 200:
            # Lấy tất cả các dòng (tr) trong trang web
            cac_dong = page.css('tr')
            so_luong_trang_nay = 0
            print(f"[DEBUG] Số lượng hàng (tr) tìm thấy: {len(cac_dong)}")

            for dong in cac_dong:
                # 1. Tìm Gói Thầu (Vào tận thẻ a)
                a_goi_thau = dong.css('td[data-column="Gói thầu"] a')

                if a_goi_thau:
                    # --- LẤY TÊN VÀ MÃ GÓI THẦU CHUẨN XÁC ---
                    # Lấy tên gói thầu từ title
                    goi_thau = a_goi_thau[0].attrib.get('title', '').strip()
                    if not goi_thau:
                        goi_thau = a_goi_thau[0].text.strip()

                    # Lấy Mã TBMT
                    ma_tbmt_node = a_goi_thau[0].css('span.bidding-code')
                    ma_tbmt = _extract_ma_tbmt(
                        _extract_text(ma_tbmt_node[0]) if ma_tbmt_node else "",
                        a_goi_thau[0].attrib.get('title', ''),
                        a_goi_thau[0].text,
                    )

                    # Lấy link gốc
                    link = a_goi_thau[0].attrib.get('href', '')
                    full_link = link if link.startswith('http') else "https://dauthau.asia" + link

                    if not goi_thau:
                        goi_thau = full_link.split('/')[-1]  # Phương án back-up cuối cùng

                    # 2. Ngày đăng tải
                    div_ngay_dang = dong.css('td[data-column="Ngày đăng tải"] div')
                    ngay_dang = div_ngay_dang[0].text.strip() if div_ngay_dang else ""

                    # 3. Đóng thầu (ở danh sách)
                    div_dong_thau = dong.css('td[data-column="Đóng thầu"] div')
                    dong_thau = _extract_text(div_dong_thau[0]) if div_dong_thau else ""

                    # Lọc chỉ theo năm, không lấy ngày tháng
                    nam_hop_le = False
                    try:
                        # Lấy 4 ký tự cuối cùng (năm) từ ngày đăng tải
                        if len(ngay_dang) >= 4:
                            nam = int(ngay_dang[-4:])
                            if nam in [2025, 2026]:
                                nam_hop_le = True
                                if "số hóa" in goi_thau.lower():
                                    # Lưu từng bản ghi phù hợp ra file riêng biệt
                                    record = {
                                        "Mã TBMT": ma_tbmt,
                                        "Gói thầu": goi_thau,
                                        "Ngày đăng tải": ngay_dang,
                                        "Đóng thầu": dong_thau,
                                        "Đường dẫn": full_link
                                    }
                                    filtered_links.append(record)
                                    so_luong_trang_nay += 1
                                    # Log chi tiết từng gói thầu trên trang
                                    print(f"[CHECK] Gói thầu: {goi_thau} | Ngày đăng: {ngay_dang}")
                    except Exception as e:
                        pass

            print(f"[+] Hoàn thành Trang {so_trang}: Thu được {so_luong_trang_nay} gói thầu phù hợp.")

        else:
            print(f"[!] Lỗi ở Trang {so_trang} (Mã: {page.status}). Dừng Crawler!")
            break

    # === LẤY CHI TIẾT TỪ CÁC LINK ĐÃ LỌC ===
    all_data = []
    for item in filtered_links:
        link = item["Đường dẫn"]
        print(f"\n---> Đang lấy chi tiết: {link}")
        try:
            detail_page = fetcher.fetch(link)
        except Exception as ex:
            print(f"[!] Timeout/lỗi khi lấy chi tiết, bỏ qua link: {link} | {ex}")
            time.sleep(2)
            continue

        if detail_page.status == 200:
            details = _parse_detail_fields(detail_page, fetcher=fetcher, detail_url=link)

            # Đồng bộ một số trường cốt lõi giữa list và detail.
            detail_ma = _extract_ma_tbmt(details.get("Mã TBMT", ""), item.get("Mã TBMT", ""))
            if detail_ma:
                item["Mã TBMT"] = detail_ma

            if not item.get("Đóng thầu"):
                item["Đóng thầu"] = details.get("Thời điểm đóng thầu", "") or details.get("Thời điểm đng thầu", "")

            all_data.append({**item, **details})
        else:
            print(f"[!] Lỗi lấy chi tiết: {link} | Status: {detail_page.status}")
        time.sleep(0)  # Nghỉ để tránh bị block

    # === TỔNG KẾT VÀ LƯU FILE EXCEL THEO TEMPLATE ===
    if all_data:
        print(f"\n[DEBUG] Tổng số bản ghi: {len(all_data)}")
        file_name = export_to_template_workbook(all_data)
        print("\n===============================")
        print(f"Đã thu thập thành công tổng cộng {len(all_data)} gói thầu số hóa.")
        print(f"Dữ liệu đã được lưu vào: {file_name}")
        print("===============================\n")
    else:
        print("\n[!] Không lấy được dữ liệu phù hợp.")

    return filtered_links

def crawl_detail_dauthau(url):
    fetcher = StealthyFetcher(headless=True)
    page = fetcher.fetch(url)
    if page.status == 200:
        details = _parse_detail_fields(page, fetcher=fetcher, detail_url=url)
        return details
    else:
        return None


def export_to_template_workbook(all_data):
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Không tìm thấy template: {TEMPLATE_PATH}")

    wb = load_workbook(TEMPLATE_PATH)
    if "DanhSachGoiThau" not in wb.sheetnames:
        raise ValueError("Template phải có sheet 'DanhSachGoiThau'.")

    detail_sheet_name = "Detail" if "Detail" in wb.sheetnames else "STT" if "STT" in wb.sheetnames else None
    if not detail_sheet_name:
        raise ValueError("Template phải có sheet detail: 'Detail' (mới) hoặc 'STT' (cũ).")

    ws_list = wb["DanhSachGoiThau"]
    ws_detail_template = wb[detail_sheet_name]

    list_headers = {}
    list_headers_norm = {}
    for col in range(1, ws_list.max_column + 1):
        header = ws_list.cell(1, col).value
        if header:
            header_text = str(header).strip()
            list_headers[header_text] = col
            list_headers_norm[_normalize_key(header_text)] = col

    def _get_list_col(*candidates):
        for c in candidates:
            col = _lookup_col(list_headers, list_headers_norm, c, LIST_HEADER_ALIASES.get(c, []))
            if col:
                return col
        return None

    stt_col = _get_list_col("STT")
    ma_tbmt_col = _get_list_col("Mã TBMT")
    goi_thau_col = _get_list_col("Gói thầu")
    ngay_dang_col = _get_list_col("Ngày đăng tải")
    duong_dan_col = _get_list_col("Đường dẫn")
    dong_thau_col = _get_list_col("Đóng thầu", "Thời điểm đóng thầu")

    # Lập map cột của sheet detail dựa trên hàng tiêu đề (row 1).
    detail_header_col_map = {}
    detail_header_col_map_norm = {}
    for col in range(1, ws_detail_template.max_column + 1):
        header = ws_detail_template.cell(1, col).value
        if header:
            header_text = str(header).strip()
            detail_header_col_map[header_text] = col
            detail_header_col_map_norm[_normalize_key(header_text)] = col

    # Xoa du lieu cu tren sheet detail (tu dong 2 tro di) de tranh ton du lieu lan chay truoc.
    for detail_row in range(2, ws_detail_template.max_row + 1):
        for col in detail_header_col_map.values():
            ws_detail_template.cell(detail_row, col).value = ""

    for idx, record in enumerate(all_data, start=1):
        row = idx + 1
        detail_row = idx + 1
        url = record.get("Đường dẫn", "")

        # Ghi detail cua tung ban ghi vao cung 1 sheet detail, moi ban ghi 1 dong.
        stt_detail_col = _lookup_col(detail_header_col_map, detail_header_col_map_norm, "STT", DETAIL_HEADER_ALIASES.get("STT", []))
        ngay_dang_detail_col = _lookup_col(detail_header_col_map, detail_header_col_map_norm, "Ngày đăng tải", DETAIL_HEADER_ALIASES.get("Ngày đăng tải", []))
        if stt_detail_col:
            ws_detail_template.cell(detail_row, stt_detail_col).value = idx
        if ngay_dang_detail_col:
            ws_detail_template.cell(detail_row, ngay_dang_detail_col).value = record.get("Ngày đăng tải", "")

        for key, value in record.items():
            key_text = str(key).strip()
            col = _lookup_col(
                detail_header_col_map,
                detail_header_col_map_norm,
                key_text,
                DETAIL_HEADER_ALIASES.get(key_text, []),
            )
            if col:
                ws_detail_template.cell(detail_row, col).value = value

        if stt_col:
            cell = ws_list.cell(row, stt_col)
            cell.value = idx
            cell.hyperlink = f"#'{detail_sheet_name}'!A{detail_row}"
            cell.font = Font(color="0563C1", underline="single")

        if ma_tbmt_col:
            ws_list.cell(row, ma_tbmt_col).value = record.get("Mã TBMT", "")
        if goi_thau_col:
            ws_list.cell(row, goi_thau_col).value = record.get("Gói thầu", "")
        if ngay_dang_col:
            ws_list.cell(row, ngay_dang_col).value = record.get("Ngày đăng tải", "")
        if dong_thau_col:
            ws_list.cell(row, dong_thau_col).value = record.get("Đóng thầu", "")
        if duong_dan_col:
            ws_list.cell(row, duong_dan_col).value = url


    thoi_gian_hien_tai = time.strftime("%d%m%Y_%H%M%S")
    RESULT_PATH.mkdir(parents=True, exist_ok=True)
    output_path = RESULT_PATH / f"DanhSachGoiThau_SoHoa_{thoi_gian_hien_tai}.xlsx"
    wb.save(output_path)
    return str(output_path)

if __name__ == "__main__":
    crawler_dauthau_chuyen_nghiep()
